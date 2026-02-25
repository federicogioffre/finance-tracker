"""
Parser for Fineco bank Excel exports.

Fineco's Excel layout:
  - Rows 1–N:  account metadata (skipped)
  - One row:   column headers ("Data operazione", "Entrate", "Uscite", ...)
  - Remaining: transaction data
  - Last rows: footer totals (skipped — no valid date)

Amounts use Italian notation: "1.234,56" → 1234.56
Dates use DD/MM/YYYY.
"""

from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl


# Column names as they appear in Fineco exports (case-insensitive match)
_COL_DATE = "data operazione"
_COL_INCOME = "entrate"
_COL_EXPENSE = "uscite"
_COL_DESC = "descrizione"
_COL_DESC_FULL = "descrizione completa"


def _parse_italian_decimal(value) -> Decimal | None:
    """Convert '1.234,56' or '1234,56' or 1234.56 to Decimal."""
    if value is None or str(value).strip() in ("", "-"):
        return None
    s = str(value).strip().replace(".", "").replace(",", ".")
    try:
        d = Decimal(s)
        return d if d > 0 else None
    except InvalidOperation:
        return None


def _parse_date(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_fineco_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse a Fineco Excel export and return a list of dicts:
      {date, amount, transaction_type, description}

    Raises ValueError with a descriptive message if the format is unrecognised.
    """
    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Find the header row (contains "Data operazione")
    header_row_idx = None
    col_map: dict[str, int] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [str(c).strip().lower() if c is not None else "" for c in row]
        if _COL_DATE in headers:
            header_row_idx = i
            col_map = {h: idx for idx, h in enumerate(headers)}
            break

    if header_row_idx is None:
        raise ValueError(
            "Intestazione non trovata. "
            "Assicurati di esportare il file direttamente da Fineco (Conto > Movimenti > Esporta)."
        )

    required = [_COL_DATE, _COL_INCOME, _COL_EXPENSE, _COL_DESC]
    missing = [c for c in required if c not in col_map]
    if missing:
        raise ValueError(f"Colonne mancanti nel file: {', '.join(missing)}")

    transactions = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        date_val = row[col_map[_COL_DATE]]
        date = _parse_date(date_val)
        if date is None:
            continue  # skip footer / empty rows

        income = _parse_italian_decimal(row[col_map[_COL_INCOME]])
        expense = _parse_italian_decimal(row[col_map[_COL_EXPENSE]])

        if income is None and expense is None:
            continue  # skip rows with no amount

        # Prefer full description if available
        desc_col = _COL_DESC_FULL if _COL_DESC_FULL in col_map else _COL_DESC
        description = str(row[col_map[desc_col]] or "").strip() or None

        if income is not None:
            transactions.append(
                {
                    "date": date,
                    "amount": income,
                    "transaction_type": "income",
                    "description": description,
                }
            )
        else:
            transactions.append(
                {
                    "date": date,
                    "amount": expense,
                    "transaction_type": "expense",
                    "description": description,
                }
            )

    return transactions
